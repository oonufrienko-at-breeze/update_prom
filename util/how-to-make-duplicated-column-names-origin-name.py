
# When pandas' read_excel function opens file it gives new uniq column name if they duplicated.
# Following code helps to restore original column name as they are in the source file.  

import pandas as pd
from openpyxl import load_workbook

# Read the Excel file
df_prom = pd.read_excel('./data/test.xlsx', engine='openpyxl')

# Load the workbook using openpyxl
book = load_workbook('./data/test.xlsx')

# Get the original column names from the first sheet
original_column_names = []
sheet = book.worksheets[0]  # Assuming the first sheet is of interest
for col in sheet.iter_cols(max_row=1, values_only=True):
    original_column_names.extend(col)

# Assign the original column names to the DataFrame
df_prom.columns = original_column_names

# Print the original column names
print(df_prom.columns)
