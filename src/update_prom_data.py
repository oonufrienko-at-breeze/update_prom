# @authors o.onufrienko@breeze.ua and AI(aka ChatGPT)

import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from datetime import datetime
import sys

print("Добрий день, Everybody\n")
# Check if the Excel files exist in the data folder
products_file_path = './data/products.xls'
tovar_file_path = './data/tovar.xls'
prom_file_path = './data/prom.xlsx'

if not os.path.isfile(products_file_path) and not os.path.isfile(tovar_file_path):
    print("Excel файли з 1С відсутні. Перевірте їх наявність у папці data. Вони мають називатись 'products.xls' та/або 'tovar.xls' \n")
    input("Для закриття вікна натисніть Enter...")
    exit()

if not os.path.isfile(prom_file_path):
    print("Excel файл з Prom.ua відсутній. Перевірте його наявність у папці data. Він має називатись 'prom.xlsx' \n")
    input("Для закриття вікна натисніть Enter...")
    exit()

# Read the 'prom' Excel file
df1 = pd.read_excel(prom_file_path)

# Read the 'products' Excel file if it exists
if os.path.isfile(products_file_path):
    df2 = pd.read_excel(products_file_path)
else:
    df2 = pd.DataFrame()  # Empty DataFrame

# Read the 'tovar' Excel file if it exists
if os.path.isfile(tovar_file_path):
    df3 = pd.read_excel(tovar_file_path)
else:
    df3 = pd.DataFrame()  # Empty DataFrame

# Define the key column names
key_column_df1 = 'Код_товару'
key_column_df23 = 3  # 3 - "Номенклатура.Код" у файлах 'products.xls' та 'tovar.xls'

# Define the fields to be read from the 'products.xls' and 'tovar.xls' files
fields_to_read = [4, 2]
# 4 - Ціна
# 2 - Залишок

# Remove all the rows with empty key values from the dataframe contains 'prom.xlsx' data
df1 = df1.dropna(subset=[key_column_df1])
# Reset indexing
df1 = df1.reset_index(drop=True)

    # Normalize key columns to a consistent numeric type (Int64) and drop invalid keys
def _normalize_key_series(series):
    numeric = pd.to_numeric(series, errors='coerce')
    return numeric.astype('Int64')

# Normalize df1 key
if key_column_df1 in df1.columns:
    df1[key_column_df1] = _normalize_key_series(df1[key_column_df1])
    df1 = df1.dropna(subset=[key_column_df1]).reset_index(drop=True)

# Normalize df2 key
if not df2.empty and key_column_df23 < len(df2.columns):
    df2.iloc[:, key_column_df23] = _normalize_key_series(df2.iloc[:, key_column_df23])
    df2 = df2.dropna(subset=[df2.columns[key_column_df23]]).reset_index(drop=True)
try:
    # Merge the data based on the key columns from products.xls if the file exists and key column is present
    if not df2.empty and key_column_df23 < len(df2.columns):
        merged_df_products = pd.merge(
            df1,
            df2.iloc[:, fields_to_read + [key_column_df23]],
            left_on=key_column_df1,
            right_on=df2.columns[key_column_df23],
            how='left'
        )
    else:
        merged_df_products = df1.copy()  # No merge, use a copy of df1

    print("Файл products.xls опрацьован\n")

except ValueError as e:
    error_message = f"Трапилася помилка формату даних! \n {e} \n Перевірте формат збереження даних з 1C. Він має бути 'Excel 97-2003 .xls' або 'Аркуш .xls' \n "
    print(error_message)
    input("Натисніть Enter щоб закрити це вікно...")  # Wait for user input before exiting or continuing
    sys.exit()  # Exit the program


    # Update the first file with the retrieved data and conditional values from products.xls if the file exists and key column is present
if not df2.empty and len(df2.columns) > key_column_df23:
    for i, row in merged_df_products.iterrows():
        zalyshok = row[merged_df_products.columns[len(merged_df_products.columns) - 2]]
        if pd.notna(zalyshok):
            df1.at[i, 'Кількість'] = zalyshok
            df1.at[i, 'Наявність'] = '!'
            df1.at[i, 'Ціна'] = row[merged_df_products.columns[len(merged_df_products.columns) - 3]]  # Assuming the 'Ціна' column is the third-to-last column
        elif pd.isna(zalyshok) and pd.notna(row[merged_df_products.columns[len(merged_df_products.columns) - 3]]):  # Assuming the 'Ціна' column is the third-to-last column
            df1.at[i, 'Кількість'] = np.nan
            df1.at[i, 'Наявність'] = '-'
            df1.at[i, 'Ціна'] = row[merged_df_products.columns[len(merged_df_products.columns) - 3]]  # Assuming the 'Ціна' column is the third-to-last column

    # Normalize df3 key
if not df3.empty and key_column_df23 < len(df3.columns):
    df3.iloc[:, key_column_df23] = _normalize_key_series(df3.iloc[:, key_column_df23])
    df3 = df3.dropna(subset=[df3.columns[key_column_df23]]).reset_index(drop=True)
try:
    # Merge the data based on the key columns from tovar.xls if the file exists and key column is present
    if not df3.empty and key_column_df23 < len(df3.columns):
        merged_df_tovar = pd.merge(
            df1,
            df3.iloc[:, fields_to_read + [key_column_df23]],
            left_on=key_column_df1,
            right_on=df3.columns[key_column_df23],
            how='left'
        )
    else:
        merged_df_tovar = df1.copy()  # No merge, use a copy of df1

    print("Файл tovar.xls опрацьован\n")

except ValueError as e:
    error_message = f"Трапилася помилка формату даних! \n '{e}' \n Перевірте формат збереження даних з 1C. Він має бути 'Excel 97-2003 .xls' або 'Аркуш .xls' \n "
    print(error_message)
    input("Натисніть Enter щоб закрити це вікно...")  # Wait for user input before exiting or continuing
    sys.exit()  # Exit the program


# Update the first file with the retrieved data and conditional values from tovar.xls if the file exists and key column is present
if not df3.empty and len(df3.columns) > key_column_df23:
    for i, row in merged_df_tovar.iterrows():
        zalyshok = row[merged_df_tovar.columns[len(merged_df_tovar.columns) - 2]]
        if pd.notna(zalyshok):
            df1.at[i, 'Кількість'] = zalyshok
            df1.at[i, 'Наявність'] = '!'
            df1.at[i, 'Ціна'] = row[merged_df_tovar.columns[len(merged_df_tovar.columns) - 3]]  # Assuming the 'Ціна' column is the third-to-last column
        elif pd.isna(zalyshok) and pd.notna(row[merged_df_tovar.columns[len(merged_df_tovar.columns) - 3]]):  # Assuming the 'Ціна' column is the third-to-last column
            df1.at[i, 'Кількість'] = np.nan
            df1.at[i, 'Наявність'] = '-'
            df1.at[i, 'Ціна'] = row[merged_df_tovar.columns[len(merged_df_tovar.columns) - 3]]  # Assuming the 'Ціна' column is the third-to-last column


# Generate the output file name with the date and time
now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
output_file = f'./output/prom_update_{now}.xlsx'

# Save the updated data to the new Excel file while preserving the original column names
df1.to_excel(output_file, index=False)

# Load the workbook from the "prom.xlsx" file
prom_wb = load_workbook(prom_file_path)
prom_ws = prom_wb.active

# Load the updated workbook and preserve the original column names
updated_wb = load_workbook(output_file)
updated_ws = updated_wb.active

# Update the column names in the updated workbook to match the original column names
for col_num, value in enumerate(prom_ws[1], 1):
    updated_ws.cell(row=1, column=col_num).value = value.value

# Save the updated workbook with the original column names
updated_wb.save(output_file)
input("Prom файл оновився. \nПеревірте папку output. \nДля закриття вікна натисніть Enter ...")